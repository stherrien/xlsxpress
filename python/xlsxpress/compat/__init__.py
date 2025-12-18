"""
OpenPyXL Compatibility Layer

This module provides a drop-in replacement for OpenPyXL,
allowing existing code to benefit from XlsXpress performance
with minimal changes.

Examples:
    Replace::

        from openpyxl import load_workbook

    With::

        from xlsxpress.compat import openpyxl
        load_workbook = openpyxl.load_workbook

    Or import the whole module::

        from xlsxpress.compat import openpyxl

        wb = openpyxl.load_workbook("data.xlsx")
        ws = wb.active
        ws["A1"] = "Hello"
        wb.save("output.xlsx")
"""

# TODO: Implement OpenPyXL compatibility in Phase 4

__all__ = [
    "openpyxl",
]
